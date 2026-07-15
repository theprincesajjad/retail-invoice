"""Inventory spreadsheet download / import (Excel + CSV).

Workflow:
  1. Download inventory → current products with Status = "In inventory"
  2. Add new rows (fill SKU + name + …); leave Status blank or "New"
  3. Import → only rows with a SKU that are not already in inventory are added
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from categories import get_product_categories, normalize_category
from models import Product

# Canonical headings (row 1)
INVENTORY_HEADERS = [
    "SKU",
    "Product Name",
    "Details",
    "Category",
    "Qty",
    "Price",
    "Status",
]

# Backwards-compatible alias used by older docs/tests
TEMPLATE_HEADERS = INVENTORY_HEADERS

STATUS_IN_INVENTORY = "In inventory"
STATUS_NEW = "New"

HEADER_ALIASES = {
    "sku": "SKU",
    "code": "SKU",
    "product sku": "SKU",
    "product code": "SKU",
    "product name": "Product Name",
    "name": "Product Name",
    "item": "Product Name",
    "item name": "Product Name",
    "details": "Details",
    "detail": "Details",
    "serial": "Details",
    "serial number": "Details",
    "s/n": "Details",
    "description": "Details",
    "category": "Category",
    "catagory": "Category",  # common misspelling
    "categories": "Category",
    "type": "Category",
    "qty": "Qty",
    "quantity": "Qty",
    "stock": "Qty",
    "in stock": "Qty",
    "price": "Price",
    "unit price": "Price",
    "cost": "Price",
    "status": "Status",
    "state": "Status",
}


class ImportResult:
    def __init__(self):
        self.added = 0
        self.updated = 0
        self.skipped = 0
        self.errors: list[str] = []

    @property
    def ok_count(self) -> int:
        return self.added + self.updated


def _normalize_header(raw: str) -> str | None:
    key = (raw or "").strip().lower()
    return HEADER_ALIASES.get(key)


def _map_headers(raw_headers: Iterable[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        canon = _normalize_header(str(h) if h is not None else "")
        if canon and canon not in mapping:
            mapping[canon] = i
    return mapping


def _cell(row: list, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    val = row[index]
    if val is None:
        return ""
    return str(val).strip()


def _parse_price(raw: str) -> float:
    clean = raw.replace("$", "").replace(",", "").strip()
    if not clean:
        return 0.0
    return float(clean)


def _parse_qty(raw: str) -> int:
    clean = raw.replace(",", "").strip()
    if not clean:
        return 0
    return int(float(clean))


def _is_in_inventory_status(status: str) -> bool:
    s = (status or "").strip().lower()
    return s in ("in inventory", "imported", "exists", "synced", "in stock system")


def _rows_from_xlsx(path: Path) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = ["" if c is None else str(c) for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    return headers, data


def _rows_from_csv(path: Path) -> tuple[list[str], list[list]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_product_rows(path: Path | str) -> list[dict]:
    """Parse a spreadsheet into product dicts. Raises ValueError on bad format."""
    path = Path(path)
    if not path.exists():
        raise ValueError(f"File not found: {path}")

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        headers, data = _rows_from_xlsx(path)
    elif suffix in (".csv", ".txt"):
        headers, data = _rows_from_csv(path)
    else:
        raise ValueError(
            "Use an Excel (.xlsx) or CSV file. Google Sheets: File → Download → Excel or CSV."
        )

    mapping = _map_headers(headers)
    if "SKU" not in mapping:
        raise ValueError(
            "Missing required column: SKU.\n"
            f"Expected headings: {', '.join(INVENTORY_HEADERS)}"
        )
    if "Product Name" not in mapping:
        raise ValueError(
            "Missing required column: Product Name.\n"
            f"Expected headings: {', '.join(INVENTORY_HEADERS)}"
        )

    products: list[dict] = []
    for row_num, row in enumerate(data, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        products.append({
            "row": row_num,
            "sku": _cell(row, mapping.get("SKU")),
            "name": _cell(row, mapping.get("Product Name")),
            "serial_number": _cell(row, mapping.get("Details")),
            "category": _cell(row, mapping.get("Category")),
            "qty": _cell(row, mapping.get("Qty")),
            "price": _cell(row, mapping.get("Price")),
            "status": _cell(row, mapping.get("Status")),
        })
    return products


def write_inventory_workbook(
    path: Path | str,
    products: list[Product] | None = None,
    *,
    blank_rows: int = 8,
) -> Path:
    """Write current inventory (Status=In inventory) plus blank rows for new items."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    products = list(products or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    for col, header in enumerate(INVENTORY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    status_fill = PatternFill("solid", fgColor="D1FAE5")
    for i, p in enumerate(products, start=2):
        values = [
            p.sku or "",
            p.name,
            p.serial_number or "",
            p.category or "Other",
            p.qty,
            p.price,
            STATUS_IN_INVENTORY,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 7:
                cell.fill = status_fill

    start_blank = len(products) + 2
    for r in range(start_blank, start_blank + blank_rows):
        ws.cell(row=r, column=7, value="")  # Status blank = new on import

    # Category dropdown for convenience
    cats = get_product_categories()
    if cats:
        try:
            joined = ",".join(c.replace(",", " ") for c in cats)
            if len(joined) < 250:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{joined}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                dv.error = "Pick a category from the list or type a new one"
                dv.errorTitle = "Category"
                ws.add_data_validation(dv)
                last_row = max(start_blank + blank_rows - 1, 2)
                dv.add(f"D2:D{last_row}")
        except Exception:
            pass
    widths = {"A": 14, "B": 32, "C": 36, "D": 16, "E": 10, "F": 12, "G": 14}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    # Notes sheet
    notes = wb.create_sheet("How to use")
    notes["A1"] = "How to keep inventory up to date"
    notes["A1"].font = Font(bold=True)
    notes["A3"] = "1. Rows marked Status = In inventory are already in Retail Invoice — Import skips them."
    notes["A4"] = "2. Add new products on blank rows below. Always fill in SKU (required) and Product Name."
    notes["A5"] = "3. Leave Status blank (or type New) for new rows."
    notes["A6"] = "4. Save the file, then in Products click Import spreadsheet — only new SKUs are added."
    notes["A7"] = f"5. Categories: {', '.join(cats)}"
    notes.column_dimensions["A"].width = 100

    wb.save(path)
    return path


def write_inventory_csv(path: Path | str, products: list[Product] | None = None, *, blank_rows: int = 8) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    products = list(products or [])
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(INVENTORY_HEADERS)
        for p in products:
            writer.writerow([
                p.sku or "",
                p.name,
                p.serial_number or "",
                p.category or "Other",
                p.qty,
                p.price,
                STATUS_IN_INVENTORY,
            ])
        for _ in range(blank_rows):
            writer.writerow(["", "", "", "", "", "", ""])
    return path


def write_excel_template(path: Path | str) -> Path:
    """Empty inventory sheet (headings + blank rows) for first-time use."""
    return write_inventory_workbook(path, products=[], blank_rows=12)


def write_csv_template(path: Path | str) -> Path:
    return write_inventory_csv(path, products=[], blank_rows=12)


def ensure_templates(assets_dir: Path | None = None) -> tuple[Path, Path]:
    from config import ASSETS_DIR

    root = Path(assets_dir) if assets_dir else ASSETS_DIR
    xlsx = write_excel_template(root / "product_import_template.xlsx")
    csv_path = write_csv_template(root / "product_import_template.csv")
    return xlsx, csv_path


def export_inventory(path: Path | str) -> Path:
    """Download current inventory with Status column filled."""
    from database import search_products

    products = search_products("")
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return write_inventory_csv(path, products)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    return write_inventory_workbook(path, products)


def import_products_from_file(path: Path | str, *, new_only: bool = True) -> ImportResult:
    """Import products. Only adds rows with a SKU that are not already in inventory.

    - Rows without SKU are skipped
    - Rows with Status = In inventory are skipped
    - Existing SKUs are skipped (never updated when new_only=True)
    """
    from database import add_product, search_products

    result = ImportResult()
    try:
        rows = read_product_rows(path)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    if not rows:
        result.errors.append("No product rows found. Download inventory, add new rows, then try again.")
        return result

    # Cache existing SKUs (case-insensitive)
    existing_skus = {
        (p.sku or "").strip().lower()
        for p in search_products("")
        if (p.sku or "").strip()
    }
    categories = get_product_categories()

    for row in rows:
        sku = (row.get("sku") or "").strip()
        if not sku:
            result.skipped += 1
            continue

        if _is_in_inventory_status(row.get("status") or ""):
            result.skipped += 1
            continue

        sku_key = sku.lower()
        if sku_key in existing_skus:
            result.skipped += 1
            continue

        name = (row.get("name") or "").strip()
        if not name:
            result.skipped += 1
            result.errors.append(f"Row {row['row']}: Product Name is required for new items")
            continue

        try:
            price = _parse_price(row.get("price") or "")
            qty = _parse_qty(row.get("qty") or "")
            if qty < 0:
                raise ValueError("Qty cannot be negative")
            if price < 0:
                raise ValueError("Price cannot be negative")
        except ValueError as e:
            result.skipped += 1
            result.errors.append(f"Row {row['row']}: {e}")
            continue

        category = normalize_category(row.get("category") or "", categories)
        product = Product(
            id=None,
            name=name,
            serial_number=(row.get("serial_number") or "").strip(),
            sku=sku,
            price=price,
            qty=qty,
            category=category,
            created_at="",
        )
        add_product(product)
        existing_skus.add(sku_key)
        result.added += 1

    if not result.ok_count and not result.errors:
        result.errors.append(
            "No new products to add. Fill SKU on blank rows (Status empty or New), then import again."
        )
    return result
